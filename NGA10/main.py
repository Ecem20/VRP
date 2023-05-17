import numpy as np
import random as rd
from openpyxl import Workbook, load_workbook
import time

wb = load_workbook("dist1.xlsx")
ws = wb.active
wb1 = load_workbook("milk1.xlsx")
ws1 = wb1.active
A = []
B = []
for x in range(1, 12):
    for y in range(1, 12):
        A.append(round(ws.cell(x, y).value))
x = np.array(A)
shape = (11, 11)
A = x.reshape(shape)

for x in range(1, 11):
    for y in range(1, 2):
        B.append(ws1.cell(x, y).value)
B = np.array(B)

def dist(fr, to):
    return A[fr][to]

def milk(fr):
    return B[fr]

X1 = [2,5,7,6,9,3,8,10,1,4] # Initial solution
p_c = 1  # Probability of crossover
p_m = 0.3  # Probability of mutation
K = 3  # For Tournament selection
pop = 1000  # Population per generation
gen = 100  # Number of generations
cap = 2500
S1 = 1000
speed = 70
k = round(speed / 60)
price_per_litre = 23  # 1 litre mazot 23 tl
km_per_litre = 5  # 1 litre mazotla 5 km yol kat edilir
tank = 400  # 400 litre mazot deposu
tk = km_per_litre*tank #2000 km

n_list = np.empty((0, len(X1)))
for i in range(int(pop)):  # Shuffles the elements in the vector n times and stores them
    rnd_sol_1 = rd.sample(X1, len(X1))
    n_list = np.vstack((n_list, rnd_sol_1))
print(n_list)

def distance(array):
    path1 = [0]
    total1 = 0  # distance
    total2 = 0  # milk
    X0 = array.copy()
    Xtemp = []
    for f in X0:
         Xtemp.append(int(f))
    print(".......................................................................................................")
    print(Xtemp)
    total_distance = 0
    for z in range(S1):
        last = path1[-1]
        next_loc = Xtemp[0]
        Xtemp.remove(next_loc)
        total1 += dist(last, next_loc)
        total2 += milk(next_loc - 1)
        if total2 <= cap and total1 <= tk:
            path1.append(next_loc)
            print("Minimum Route:", path1, "Total Distance:", total1, "current distance:",dist(last, next_loc), "current milk:", milk(next_loc - 1), ",", "Total Milk:", total2)
        else:
            Xtemp.insert(0, next_loc)
            total2 -= milk(next_loc - 1)
            total1 -= dist(last, next_loc)
            path1.append(0)
            tdk = dist(last, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk  # distance between last station to center
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            print("Average turnaround time:", round(((k * total1) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            print("...................")
        if len(Xtemp) == 0:
            path1.append(0)
            tdk = dist(next_loc, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            print("Average turnaround time:", round(((k * total1) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            print("...................")
            if len(Xtemp) == 0:
                 break
    return total_distance

Final_Best_in_Generation_X = []
Worst_Best_in_Generation_X = []
For_Plotting_the_Best = np.empty((0, len(X1) + 1))
One_Final_Guy = np.empty((0, len(X1) + 2))
One_Final_Guy_Final = []
Min_for_all_Generations_for_Mut_1 = np.empty((0, len(X1) + 1)) #fittness value
Min_for_all_Generations_for_Mut_2 = np.empty((0, len(X1) + 1))
Min_for_all_Generations_for_Mut_1_1 = np.empty((0, len(X1) + 2)) #at what generation best value was achieved
Min_for_all_Generations_for_Mut_2_2 = np.empty((0, len(X1) + 2))
Min_for_all_Generations_for_Mut_1_1_1 = np.empty((0, len(X1) + 2))
Min_for_all_Generations_for_Mut_2_2_2 = np.empty((0, len(X1) + 2))

start_time = time.time()  # start the time
Generation = 1
for i in range(gen):
    New_Population = np.empty((0, len(X1)))  # Saving the new generation
    All_in_Generation_X_1 = np.empty((0, len(X1) + 1)) #keep every generations mutant children
    All_in_Generation_X_2 = np.empty((0, len(X1) + 1))
    Min_in_Generation_X_1 = []
    Min_in_Generation_X_2 = []
    Save_Best_in_Generation_X = np.empty((0, len(X1) + 1))#keep every solution to pick the best one
    Final_Best_in_Generation_X = []
    Worst_Best_in_Generation_X = []
    print(".........................................")
    print("--> GENERATION: #", Generation)
    Family = 1
    for j in range(int(pop / 2)):  #in each family 2 parent,2 children,2 mutant children
        print("..................")
        print("--> FAMILY: #", Family)

        Parents = np.empty((0, len(X1)))#store parents
        Parent = 1
        #TOURNAMENT SELECTION
        for i in range(2): #2 because we do it for 2 parents
            indices = np.random.choice(len(n_list), size=3, replace=False)
            Warrior_1_index = indices[0]
            Warrior_2_index = indices[1]
            Warrior_3_index = indices[2]

            Warrior_1 = n_list[Warrior_1_index, :]
            Prize_Warrior_1 = distance(Warrior_1)
            Warrior_2 = n_list[Warrior_2_index, :]
            Prize_Warrior_2 = distance(Warrior_2)
            Warrior_3 = n_list[Warrior_3_index, :]
            Prize_Warrior_3 = distance(Warrior_3)

            if Prize_Warrior_1 == min(Prize_Warrior_1, Prize_Warrior_2, Prize_Warrior_3):
                Winner = Warrior_1
            elif Prize_Warrior_2 == min(Prize_Warrior_1, Prize_Warrior_2, Prize_Warrior_3):
                Winner = Warrior_2
            else:
                Winner = Warrior_3

            Parents = np.vstack((Parents, Winner))
            print("Parents",Parents)

        Parent_1 = Parents[0]
        Parent_2 = Parents[1]

        Child_1 = np.empty((0, len(X1)))
        Child_1 = [0 for i in range(len(X1))]
        print("Child_1",Child_1)

        Child_2 = np.empty((0, len(X1)))
        Child_2 = [0 for i in range(len(X1))]
        print("Child_2",Child_2)


        #CROSSOVER
        crossover_random_number = np.random.rand()
        print(crossover_random_number)
        if crossover_random_number < p_c:
            # Choose two random numbers to crossover with their locations
            #cross over from location 1 and location 2 and exchange the genes between the two parents based on index of location 1 up until location 2.
            Cr_1 = np.random.randint(0, len(X1))
            Cr_2 = np.random.randint(0, len(X1))
            while Cr_1 == Cr_2:
                Cr_2 = np.random.randint(0, len(X1))
            print("Cr_2",Cr_2)
            print("Cr_1",Cr_1)

            if Cr_1 < Cr_2:
                Cr_2 = Cr_2 + 1 #[1:3] to include 3
                New_Dep_1 = Parent_1[Cr_1:Cr_2]  #gets the location range
                New_Dep_2 = Parent_2[Cr_1:Cr_2]  #gets the location range
                #for child 2
                Child_2[Cr_1:Cr_2] = New_Dep_1[:]
                result = [i for i in Parent_2 if i not in set(Child_2)]
                x = 0
                for i in range(len(Child_2)):
                    if Child_2[i] == 0 and x < len(result):
                        Child_2[i] = result[x]
                        x += 1
                print("Child2", Child_2)

                #for child 1
                Child_1[Cr_1:Cr_2] = New_Dep_2[:]
                result = [i for i in Parent_1 if i not in set(Child_1)]
                x = 0
                for i in range(len(Child_1)):
                    if Child_1[i] == 0 and x < len(result):
                        Child_1[i] = result[x]
                        x += 1
                print("Child1", Child_1)

            else:  # The same in reverse of Cr_1 and Cr_2

                Cr_1 = Cr_1 + 1

                New_Dep_1 = Parent_1[Cr_2:Cr_1]  #gets the location range
                New_Dep_2 = Parent_2[Cr_2:Cr_1]  #gets the location range

                #for child 2
                Child_2[Cr_2:Cr_1] = New_Dep_1[:]
                result = [i for i in Parent_2 if i not in set(Child_2)]
                x = 0
                for i in range(len(Child_2)):
                    if Child_2[i] == 0 and x < len(result):
                        Child_2[i] = result[x]
                        x += 1
                print("Child2", Child_2)

                #for child 1
                Child_1[Cr_2:Cr_1] = New_Dep_2[:]
                result = [i for i in Parent_1 if i not in set(Child_1)]
                x = 0
                for i in range(len(Child_1)):
                    if Child_1[i] == 0 and x < len(result):
                        Child_1[i] = result[x]
                        x += 1
                print("Child1", Child_1)

        else:  # If random number was above p_c,means no crossover happen,but prob. of crossover is 1 so crossover always happens
            Child_1 = Parent_1 #child 1 will be same as parent 1
            Child_2 = Parent_2 #child 2 will be same as parent 2


        #MUTATION
        #For Child1
        Ran_Mut_1 = np.random.rand()  # Probablity to Mutate,to tell if we should mutate or not
        Ran_Mut_2 = np.random.randint(0, len(X1)) #random integer
        Ran_Mut_3 = np.random.randint(0, len(X1)) #random integer
        A1 = Ran_Mut_2 # random index
        A2 = Ran_Mut_3 # random index
        while A1 == A2:
            A2 = np.random.randint(0, len(X1))
        print(A1,A2)
        if Ran_Mut_1 < p_m:  # If probablity to mutate is less than p_m, then mutate
            if A1 < A2:
                M_Child_1_Pos_1 = Child_1[A1]
                M_Child_1_Pos_2 = Child_1[A2]
                A2 = A2 + 1
                Rev_1 = Child_1[:]
                Rev_2 = list(reversed(Child_1[A1:A2]))
                t = 0
                for i in range(A1, A2):
                    Rev_1[i] = Rev_2[t]  # The reversed will become instead of the original
                    t = t + 1
                Mutated_Child_1 = Rev_1
                print("Mutated_Child_1",Mutated_Child_1)

            else:
                M_Child_1_Pos_1 = Child_1[A2]
                M_Child_1_Pos_2 = Child_1[A1]
                A1 = A1 + 1
                Rev_1 = Child_1[:]
                Rev_2 = list(reversed(Child_1[A2:A1]))
                t = 0
                for i in range(A2, A1):
                    Rev_1[i] = Rev_2[t]
                    t = t + 1
                Mutated_Child_1 = Rev_1
                print("Mutated_Child_1",Mutated_Child_1)

        else:
            Mutated_Child_1 = Child_1
            print("Child_1",Child_1)


        #For Child2
        Mutated_Child_2 = []
        Ran_Mut_1 = np.random.rand()  # Probablity to Mutate
        Ran_Mut_2 = np.random.randint(0, len(X1))
        Ran_Mut_3 = np.random.randint(0, len(X1))
        A1 = Ran_Mut_2
        A2 = Ran_Mut_3
        while A1 == A2:
            A2 = np.random.randint(0, len(X1))

        if Ran_Mut_1 < p_m:  # If probablity to mutate is less than p_m, then mutate
            if A1 < A2:
                M_Child_1_Pos_1 = Child_2[A1]
                M_Child_1_Pos_2 = Child_2[A2]
                A2 = A2 + 1
                Rev_1 = Child_2[:]
                Rev_2 = list(reversed(Child_2[A1:A2]))
                t = 0
                for i in range(A1, A2):
                    Rev_1[i] = Rev_2[t]
                    t = t + 1
                Mutated_Child_2 = Rev_1
                print("Mutated_Child_2",Mutated_Child_2)

            else:
                M_Child_1_Pos_1 = Child_2[A2]
                M_Child_1_Pos_2 = Child_2[A1]
                A1 = A1 + 1
                Rev_1 = Child_2[:]
                Rev_2 = list(reversed(Child_2[A2:A1]))
                t = 0
                for i in range(A2, A1):
                    Rev_1[i] = Rev_2[t]
                    t = t + 1
                Mutated_Child_2 = Rev_1
                print("Mutated_Child_2",Mutated_Child_2)

        else:
            Mutated_Child_2 = Child_2
            print("Child_2",Child_2)

        Total_Dist_Mut_1 = distance(Mutated_Child_1)
        Total_Dist_Mut_2 = distance(Mutated_Child_2)

        MC1 = []
        print("Mutated_Child_1",Mutated_Child_1)
        for f in Mutated_Child_1:
             MC1.append(int(f))
        All_in_Generation_X_1_1_temp = np.array(MC1)[np.newaxis]
        print("All_in_Generation_X_1_1_temp",All_in_Generation_X_1_1_temp)
        arr1_str_temp = All_in_Generation_X_1_1_temp.astype(str)
        All_in_Generation_X_1_1 = np.column_stack((Total_Dist_Mut_1, All_in_Generation_X_1_1_temp))

        MC2 = []
        print("Mutated_Child_2", Mutated_Child_2)
        for f in Mutated_Child_2:
            MC2.append(int(f))
        All_in_Generation_X_2_1_temp = np.array(MC2)[np.newaxis]
        print("All_in_Generation_X_2_1_temp",All_in_Generation_X_2_1_temp)
        arr2_str_temp = All_in_Generation_X_2_1_temp.astype(str)
        All_in_Generation_X_2_1 = np.column_stack((Total_Dist_Mut_2, All_in_Generation_X_2_1_temp))

        arr1_str = All_in_Generation_X_1_1.astype(str)
        All_in_Generation_X_1 = np.vstack((All_in_Generation_X_1, arr1_str))
        print("All_in_Generation_X_1",All_in_Generation_X_1)

        arr2_str = All_in_Generation_X_2_1.astype(str)
        All_in_Generation_X_2 = np.vstack((All_in_Generation_X_2, arr2_str))
        print("All_in_Generation_X_2",All_in_Generation_X_2)

        Save_Best_in_Generation_X = np.vstack((All_in_Generation_X_1, All_in_Generation_X_2))
        print("Save_Best_in_Generation_X",Save_Best_in_Generation_X)

        New_Population = np.vstack((New_Population, arr1_str_temp, arr2_str_temp))
        print("New Population",New_Population)

        #to find the minimum for each generation
        #for mutant child 1
        t = 0
        R_1 = []
        for i in All_in_Generation_X_1:
            if (All_in_Generation_X_1[t, :1]) <= min(All_in_Generation_X_1[:, :1]):
                R_1 = All_in_Generation_X_1[t, :]
            t = t + 1
        Min_in_Generation_X_1 = R_1[np.newaxis]
        print("Min_in_Generation_X_1",Min_in_Generation_X_1)

        #to find the minimum for each generation
        #for mutant child 2
        t = 0
        R_2 = []
        for i in All_in_Generation_X_2:
            if (All_in_Generation_X_2[t, :1]) <= min(All_in_Generation_X_2[:, :1]):
                R_2 = All_in_Generation_X_2[t, :]
            t = t + 1
        Min_in_Generation_X_2 = R_2[np.newaxis]
        print("Min_in_Generation_X_2",Min_in_Generation_X_2)
        print(".........................................")

        Family = Family + 1

    t = 0
    R_Final = []

    #in order to apply elitism we need to find the best chromosome in Save_Best_in_Generation_X
    #substitute it with the worst chromosome in that list
    for i in Save_Best_in_Generation_X:
        if (Save_Best_in_Generation_X[t, :1]) <= min(Save_Best_in_Generation_X[:, :1]):
            R_Final = Save_Best_in_Generation_X[t, :]
        t = t + 1
    Final_Best_in_Generation_X = R_Final[np.newaxis]
    print("Final_Best_in_Generation_X",Final_Best_in_Generation_X)
    For_Plotting_the_Best = np.vstack((For_Plotting_the_Best, Final_Best_in_Generation_X))

    t = 0
    R_22_Final = []
    for i in Save_Best_in_Generation_X:
        if (Save_Best_in_Generation_X[t, :1]) >= max(Save_Best_in_Generation_X[:, :1]):
            R_22_Final = Save_Best_in_Generation_X[t, :]
        t = t + 1
    Worst_Best_in_Generation_X = R_22_Final[np.newaxis]
    print("Worst_Best_in_Generation_X",Worst_Best_in_Generation_X)
    print(".........................................")


    #ELITISM-the best in the generation lives
    Darwin_Guy = Final_Best_in_Generation_X[:]
    Darwin_Guy = Darwin_Guy[0:, 1:].tolist()
    Not_So_Darwin_Guy = Worst_Best_in_Generation_X[:]
    Not_So_Darwin_Guy = Not_So_Darwin_Guy[0:, 1:].tolist()
    #where does best have its place in population
    Best_1 = np.where((New_Population == Darwin_Guy).all(axis=1))
    print("Best_1",Best_1)

    #where does worst have its place in population
    Worst_1 = np.where((New_Population == Not_So_Darwin_Guy).all(axis=1))
    print("Worst_1",Worst_1)
    print(".........................................")

    #change the worst one with the best one
    New_Population[Worst_1] = Darwin_Guy

    #new population created
    n_list = New_Population
    print("n_list",n_list)
    print(".........................................")


    # stack all the mutant children top of each other to keep track of what generation they were achieved that.
    # for minimum of in each generation, keep track of what generation it was achieved.
    Min_for_all_Generations_for_Mut_1 = np.vstack((Min_for_all_Generations_for_Mut_1, Min_in_Generation_X_1))
    print("Min_for_all_Generations_for_Mut_1", Min_for_all_Generations_for_Mut_1)
    Min_for_all_Generations_for_Mut_2 = np.vstack((Min_for_all_Generations_for_Mut_2, Min_in_Generation_X_2))
    print("Min_for_all_Generations_for_Mut_2", Min_for_all_Generations_for_Mut_2)

    # add the generation number
    Min_for_all_Generations_for_Mut_1_1 = np.insert(Min_in_Generation_X_1, 0, Generation)
    print("Min_for_all_Generations_for_Mut_1_1", Min_for_all_Generations_for_Mut_1_1)
    # add the generation number
    Min_for_all_Generations_for_Mut_2_2 = np.insert(Min_in_Generation_X_2, 0, Generation)
    print("Min_for_all_Generations_for_Mut_2_2", Min_for_all_Generations_for_Mut_2_2)

    Min_for_all_Generations_for_Mut_1_1_1 = np.vstack((Min_for_all_Generations_for_Mut_1_1_1, Min_for_all_Generations_for_Mut_1_1))
    print("Min_for_all_Generations_for_Mut_1_1_1", Min_for_all_Generations_for_Mut_1_1_1)
    Min_for_all_Generations_for_Mut_2_2_2 = np.vstack((Min_for_all_Generations_for_Mut_2_2_2, Min_for_all_Generations_for_Mut_2_2))
    print("Min_for_all_Generations_for_Mut_2_2_2", Min_for_all_Generations_for_Mut_2_2_2)

    Generation = Generation + 1

One_Final_Guy = np.vstack((Min_for_all_Generations_for_Mut_1_1_1, Min_for_all_Generations_for_Mut_2_2_2))
print("One_Final_Guy", One_Final_Guy)
print(".........................................")
t = 0
Final_Here = []
for i in One_Final_Guy:
    if (One_Final_Guy[t, 1]) <= min(One_Final_Guy[:, 1]):
        Final_Here = One_Final_Guy[t, :]
    t = t + 1
One_Final_Guy_Final = Final_Here[np.newaxis]
print("Final Result",One_Final_Guy_Final)
print("Min in all Generations:", One_Final_Guy_Final[:, 2:])
print(distance(min(One_Final_Guy_Final[:, 2:])))
print(".........................................")
print("The Lowest Cost is:",min(One_Final_Guy[:, 1]))
print("At Generation:", min(One_Final_Guy_Final[:, 0]))
comp_time = time.time() - start_time  # Keeps the difference between the end time and the start time
print(f"-> Computational Time: {comp_time} seconds")