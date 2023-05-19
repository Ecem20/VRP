import numpy as np
from openpyxl import Workbook, load_workbook
import time

wb = load_workbook("dist1.xlsx")
ws = wb.active
wb1 = load_workbook("milk1.xlsx")
ws1 = wb1.active
A = []
B = []
C = []
for x in range(1, 12):
    for y in range(1, 12):
        A.append(round(ws.cell(x, y).value))
x = np.array(A)
shape = (11, 11)
A = x.reshape(shape)


for x in range(1, 11):
    for y in range(1, 2):
        B.append(ws1.cell(x, y).value)
B = np.array(B)

def dist(fr, to):
    return A[fr][to]

def milk(fr):
    return B[fr]

cap = 2500
speed = 70
k = round(speed / 60)
price_per_litre = 23  # 1 litre mazot 23 tl
km_per_litre = 5  # 1 litre mazot 5 km yol kat edilir
tank = 400  # 400 litre mazot deposu
tk = km_per_litre*tank #2000 km
S = 1000

start_time = time.time()  # Keeps the start time
path = [0]
total_Distance = 0
totalMilk = 0
mask = [True] *len(B)
for i in range(S):
    last = path[-1]
    max_val = -1
    for i in range(len(mask)):
        if mask[i] and B[i] > max_val:
            max_val = B[i]
            next_loc = i
    mask[next_loc] = False
    total_Distance += dist(last, next_loc+1)
    totalMilk += milk(next_loc)
    if (tk - total_Distance >= 0 and cap-totalMilk >= 0):
        path.append(next_loc+1)
        print("Minimum Route:", path, "Total Distance:", total_Distance, "current distance:",dist(last,next_loc+1),"current milk:",milk(next_loc), "Total Milk:", totalMilk)
    else:
        mask[next_loc] = True
        totalMilk -= milk(next_loc)
        total_Distance -= dist(last, next_loc+1)

        max_val = -1
        next_loc = -1
        for i in range(len(B)):
            if mask[i] and B[i] > max_val and B[i] <= cap - totalMilk:
                max_val = B[i]
                next_loc = i

        if next_loc != -1: #founds a next location
            mask[next_loc] = False
            total_Distance += dist(last, next_loc + 1)
            totalMilk += milk(next_loc)
            path.append(next_loc + 1)
            print("Minimum Route:", path, "Total Distance:", total_Distance, "current distance:",dist(last, next_loc + 1), "current milk:", milk(next_loc), "Total Milk:", totalMilk)
            last = next_loc + 1

        path.append(0)
        tdk = dist(last, 0)
        total_Distance += tdk
        for i in range(S):
            if (tk - total_Distance < 0):
                total_Distance -= dist(path[len(path) - 2], 0)
                total_Distance -= dist(path[len(path) - 3], path[len(path) - 2])
                tdk = dist(path[len(path) - 3], 0)
                total_Distance += tdk #distance between last station to center
                totalMilk -= milk(path[len(path) - 2]-1)
                mask[path[len(path) - 2]-1] = True
                path.remove(path[len(path) - 2])
        print("Minimum Route:", path,"Total Distance:", total_Distance,"km.","Total Milk:",totalMilk)
        print("Average turnaround time:", round(((k * total_Distance) /60)), "hr.")
        print("Total amount of fuel consumed:", round(total_Distance / km_per_litre), "litre")
        print("Total cost:",round(total_Distance / km_per_litre) * price_per_litre, "tl")
        total_Distance = 0
        totalMilk = 0
        path = [0]
        print("...................")

    all_false = True
    for i in mask:
        if i:
            all_false = False
            break
    if all_false:
        path.append(0)
        tdk = dist(next_loc+1, 0)
        total_Distance += tdk
        for i in range(S):
            if (tk - total_Distance < 0):
                total_Distance -= dist(path[len(path) - 2], 0)
                total_Distance -= dist(path[len(path) - 3], path[len(path) - 2])
                tdk = dist(path[len(path) - 3], 0)
                total_Distance += tdk
                totalMilk -= milk(path[len(path) - 2] - 1)
                mask[path[len(path) - 2]-1] = True
                path.remove(path[len(path) - 2])
        print("Minimum Route:", path, "Total Distance:", total_Distance, "km.", "Total Milk:", totalMilk)
        print("Average turnaround time:", round(((k * total_Distance) / 60)), "hr.")
        print("Total amount of fuel consumed:", round(total_Distance / km_per_litre), "litre")
        print("Total cost:", round(total_Distance / km_per_litre) * price_per_litre, "tl")
        total_Distance = 0
        totalMilk = 0
        path = [0]
        print("...................")
        for i in mask:
            if i:
                all_false = False
                break
        if all_false:
            break

comp_time = time.time() - start_time  # Keeps the difference between the end time and the start time
print(f"-> Computational Time: {comp_time} seconds")  # Prints Computational Time

