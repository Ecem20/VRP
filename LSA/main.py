import numpy as np
from openpyxl import Workbook, load_workbook
import time
import random as rd

wb = load_workbook("dist1.xlsx")
ws = wb.active
wb1 = load_workbook("milk1.xlsx")
ws1 = wb1.active
A = []
B = []
s_2 = ws.max_column
k_1 = ws1.max_row
k_2 = ws1.max_column

for x in range(1, s_2 + 1):
    for y in range(1, s_2 + 1):
        A.append(round(ws.cell(x, y).value))
A = np.array(A).reshape(s_2, s_2)
print(A)
for x in range(1, k_1+1):
    for y in range(1, k_2 + 1):
        B.append(ws1.cell(x, y).value)
B = np.array(B)
print(B)
def dist(fr, to):
    return A[fr][to]

def milk(fr):
    return B[fr]

cap = 2500
X1 = [8,2,7,6,3,9,4,1,10,5] # Initial solution

M = 50000
speed = 70
k = round(speed / 60)
price_per_litre = 23  # 1 litre mazot 23 tl
km_per_litre = 5  # 1 litre mazotla 5 km yol kat edilir
tank = 400  # 400 litre mazot deposu
tk = km_per_litre*tank #2000 km
S1 = 1000
T0 = 3000  # initial temperature
Alpha = 0.85

def distance(array):
    path1 = [0]
    total1 = 0  # distance
    total2 = 0  # milk
    Xtemp = array.copy()
    print(Xtemp)
    total_distance = 0
    for z in range(S1):
        last = path1[-1]
        next_loc = Xtemp[0]
        Xtemp.remove(next_loc)
        total1 += dist(last, next_loc)
        total2 += milk(next_loc - 1)
        if total2 <= cap and total1 <= tk:
            path1.append(next_loc)
            print("Minimum Route:", path1, "Total Distance:", total1, "current distance:",dist(last, next_loc), "current milk:", milk(next_loc - 1), ",", "Total Milk:", total2)
        else:
            Xtemp.insert(0, next_loc)
            total2 -= milk(next_loc - 1)
            total1 -= dist(last, next_loc)
            path1.append(0)
            tdk = dist(last, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk  # distance between last station to center
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            print("Average turnaround time:", round(((k * total1) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            print("...................")
        if len(Xtemp) == 0:
            path1.append(0)
            tdk = dist(next_loc, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            print("Average turnaround time:", round(((k * total1) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            print("...................")
            if len(Xtemp) == 0:
                 break
    return total_distance



start_time = time.time()  # start the time
for i in range(M):
        # select 2 random integer to swap
        print("previous route:", X1)
        selected_values = rd.sample(X1, k=2)
        A1 = selected_values[0]
        A2 = selected_values[1]
        print(A1,A2)

        #swap selected random values
        Xtempp = []
        for x in X1:
            if x == A1:
                Xtempp.append(A2)
            elif x == A2:
                Xtempp.append(A1)
            else:
                Xtempp.append(x)
        print("current route",Xtempp)

        # for previous route
        prev_route_distance = distance(X1)
        # for current route
        current_route_distance = distance(Xtempp)
        print(X1,"prev_route_distance", prev_route_distance)
        print(Xtempp,"current_route_distance", current_route_distance)

        rand_num = rd.random()
        print("random number:",rand_num)
        formul = 1 / (np.exp((current_route_distance - prev_route_distance) / T0))  # The formula to accept moves
        print("formul",formul)

        if current_route_distance <= prev_route_distance:  # if current distance is better
             X1 = Xtempp
             prev_route_distance = current_route_distance
        elif rand_num <= formul:  # If random number is less than the formula
             X1 = Xtempp
             prev_route_distance = current_route_distance
        else:  # Don't accept the potential solution and stay where you are
             X1 = X1
             prev_route_distance = prev_route_distance

        T0 = T0 /(1+(Alpha*i))  # Decrease the temp.
        print("..........")
        ind = i

print("Final Solution is: ", X1)
print("Minimized Distance at Final Solution is: ", prev_route_distance)
T0 = T0 * (1 + (Alpha * ind))
print("T0:",T0)
comp_time = time.time() - start_time  # Keeps the difference between the end time and the start time
print(f"-> Computational Time: {comp_time} seconds")  # Prints Computational Time