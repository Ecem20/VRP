import numpy as np
from gurobipy import *
from openpyxl import load_workbook

wb = load_workbook("dist1.xlsx")
ws = wb.active
wb1 = load_workbook("milk.xlsx")
ws1 = wb1.active
A = []
demand = []
s_2 = ws.max_column
k_1 = ws1.max_row
k_2 = ws1.max_column

for x in range(1, s_2 + 1):
    for y in range(1, s_2 + 1):
        A.append(round(ws.cell(x, y).value))
A = np.array(A).reshape(s_2, s_2)
for x in range(1, k_1+1):
    for y in range(1, k_2 + 1):
        demand.append(ws1.cell(x, y).value)
demand = np.array(demand)

distance = A
nodes = range(distance.shape[0])
sum_waste = np.sum(demand) #sum of the demands
Q = 2500 #capacity of vehicle
max_distance = 2000 #maximum distance for each route
routes = math.ceil((sum_waste/Q)) #for minimum route number
mdl = Model('CVRP') #Create a new model

#Create variables
x = mdl.addVars(nodes,nodes,lb=0,vtype=GRB.BINARY,name='x') #location
u = mdl.addVars(nodes,lb=0,vtype=GRB.INTEGER,name='u')

#constraints
mdl.addConstrs(((quicksum(x[i,j] for j in nodes if j != i) == 1) for i in nodes if i != 0), name = 'constraint 1') #each node (except node 0) is connected to exactly one other node.
mdl.addConstrs(((quicksum(x[i,j] for i in nodes if i != j) == 1) for j in nodes if j != 0), name = 'constraint 2') #each node (except node 0) is connected from exactly one other node
mdl.addConstr(((quicksum(x[0,j] for j in nodes) == routes)), name = 'constraint 3') #The number of arcs coming out of zero is equal to the number of routes.
mdl.addConstr(((quicksum(x[i,0] for i in nodes) == routes)), name = 'constraint 4') #The number of arcs coming to zero is equal to the number of routes.
mdl.addConstrs(((u[i] - u[j] + Q*x[i,j] <= Q - demand[j]) for i in nodes for j in nodes if i!=j and i != 0 and j != 0), name = 'constraint 5') #subtour elimination
mdl.addConstrs(((u[i] <= Q) for i in nodes), name = 'constraint 6') #capacity constraint
mdl.addConstrs(((u[i] >= demand[i]) for i in nodes), name = 'constraint 7') #subtour elimination
mdl.addConstrs((u[i] - u[j] + distance[i][j] * x[i,j] <= max_distance for i in nodes for j in nodes if i != j and i != 0 and j != 0), name='constraint 8') #distance constraint

#objective function
mdl.setObjective((quicksum(distance[i][j]*x[i,j] for i in nodes for j in nodes if i != j)),GRB.MINIMIZE) #objective function,find minimum distance
mdl.update()
mdl.optimize()

object_Value = mdl.objVal
print("Objective value is: ", object_Value)
active_arcs=[v for v in mdl.getVars() if v.x!=0]
print(active_arcs)