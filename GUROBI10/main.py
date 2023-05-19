import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from gurobipy import *

d = pd.read_excel('dist.xlsx')
nodes = range(d.shape[0])
distance = [[d[j][i] for j in nodes] for i in nodes]
waste = pd.read_excel('milk.xlsx')
waste = np.array(waste)
sum_waste = np.sum(waste)
Q = 2500 #capacity of vehicle
max_distance = 2000 #maximum distance for each route
routes = math.ceil((sum_waste/Q))
#Create a new model
mdl=Model('CVRP')

#Create variables
x = mdl.addVars(nodes,nodes,lb=0,vtype=GRB.BINARY,name='x') #station
u = mdl.addVars(nodes,lb=0,vtype=GRB.INTEGER,name='u') #waste

#constraints
mdl.addConstrs(((quicksum(x[i,j] for j in nodes if j != i) == 1) for i in nodes if i != 0), name = 'constraint 1') #each node (except node 0) is connected to exactly one other node.
mdl.addConstrs(((quicksum(x[i,j] for i in nodes if i != j) == 1) for j in nodes if j != 0), name = 'constraint 2') #each node (except node 0) is connected from exactly one other node
mdl.addConstr(((quicksum(x[0,j] for j in nodes) == routes)), name = 'constraint 3') #The number of arcs coming out of zero is equal to the number of routes.
mdl.addConstr(((quicksum(x[i,0] for i in nodes) == routes)), name = 'constraint 4') #The number of arcs coming to zero is equal to the number of routes.
mdl.addConstrs(((u[i] - u[j] + Q*x[i,j] <= Q - waste[j]) for i in nodes for j in nodes if i!=j and i != 0 and j != 0), name = 'constraint 5') #subtour elimination
mdl.addConstrs(((u[i] <= Q) for i in nodes), name = 'constraint 6') #capacity constraint
mdl.addConstrs(((u[i] >= waste[i]) for i in nodes), name = 'constraint 7') #subtour elimination
mdl.addConstrs((u[i] - u[j] + distance[i][j] * x[i,j] <= max_distance for i in nodes for j in nodes if i != j and i != 0 and j != 0), name='constraint 8') #distance constraint

#objective function
mdl.setObjective((quicksum(distance[i][j]*x[i,j] for i in nodes for j in nodes if i != j)),GRB.MINIMIZE)
mdl.update()
mdl.optimize()

object_Value = mdl.objVal
print("Objective value is: ", object_Value)

wb = load_workbook("dist11.xlsx")
ws = wb.active
wb1 = load_workbook("milk11.xlsx")
ws1 = wb1.active
A = []
B = []
for d in range(1, 12):
    for y in range(1, 12):
        A.append(round(ws.cell(d, y).value))
d = np.array(A)
shape = (11, 11)
A = d.reshape(shape)

for d in range(1, 11):
    for y in range(1, 2):
        B.append(ws1.cell(d, y).value)
B = np.array(B)

def dist(fr, to):
    return A[fr][to]

def milk(fr):
    return B[fr]

speed = 70
kl = round(speed / 60)
price_per_litre = 23  # 1 litre mazot 23 tl
km_per_litre = 5  # 1 litre mazotla 5 km yol kat edilir
tank = 400  # 400 litre mazot deposu
tk = km_per_litre*tank #2000 km
S1 = 1000

def distance(array):
    path1 = [0]
    total1 = 0  # distance
    total2 = 0  # milk
    X0 = array.copy()
    Xtemp = []
    for f in X0:
         Xtemp.append(int(f))
    print(".......................................................................................................")
    print(Xtemp)
    total_distance = 0
    for z in range(S1):
        last = path1[-1]
        next_loc = Xtemp[0]
        Xtemp.remove(next_loc)
        total1 += dist(last, next_loc)
        total2 += milk(next_loc - 1)
        if total2 <= Q and total1 <= tk:
            path1.append(next_loc)
            print("Minimum Route:", path1, "Total Distance:", total1, "current distance:",dist(last, next_loc), "current milk:", milk(next_loc - 1), ",", "Total Milk:", total2)
        else:
            Xtemp.insert(0, next_loc)
            total2 -= milk(next_loc - 1)
            total1 -= dist(last, next_loc)
            path1.append(0)
            tdk = dist(last, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk  # distance between last station to center
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            print("Average turnaround time:", round(((kl * total1) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            print("...................")
        if len(Xtemp) == 0:
            path1.append(0)
            tdk = dist(next_loc, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            print("Average turnaround time:", round(((kl * total1) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            print("...................")
            if len(Xtemp) == 0:
                 break
    return total_distance


vis = []
Sol_x = np.zeros([len(nodes), len(nodes)])
for i in nodes:
    for j in nodes:
        Sol_x[i, j] = x[i, j].getAttr("X")
        if Sol_x[i, j]:
            vis.append((i, j))
visited = np.array(vis)
if visited[0][0] == 0:
    sol = [visited[0][0], visited[0][1]]
visited = np.delete(visited, 0, axis=0)
solution = []
for i in visited:
        next_ind = int(np.where(visited[:, 0] == sol[-1])[0])  #select whichever pair's last position is the same as the first position of the remaining pair
        sol.append(visited[next_ind][1])
        visited = np.delete(visited, next_ind, axis=0)
        if sol[0] == sol[-1]:
            sol = np.asarray(sol)
            solution.append(sol)
            sol = sol[1:]
            sol = sol[:-1]
            distance(sol)  #route
            used = []
            for j in solution:
                for k in j:
                    used.append(k)
            remain = list(set(nodes) - set(used))
            if remain == []:
                break
            sol = [visited[0][0], visited[0][1]]