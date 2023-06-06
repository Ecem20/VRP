import numpy as np
import random as rd
from openpyxl import load_workbook
import time

wb = load_workbook("dist1.xlsx")
ws = wb.active
wb1 = load_workbook("milk1.xlsx")
ws1 = wb1.active
A = []
B = []
for x in range(1, 12):
    for y in range(1, 12):
        A.append(round(ws.cell(x, y).value))
x = np.array(A)
shape = (11, 11)
A = x.reshape(shape)

for x in range(1, 11):
    for y in range(1, 2):
        B.append(ws1.cell(x, y).value)
B = np.array(B)

def dist(fr, to):
    return A[fr][to]

def milk(fr):
    return B[fr]

X1 = [8,2,7,6,3,9,4,1,10,5] # Initial solution
p_c = 1  # Probability of crossover
p_m = 0.3  # Probability of mutation
K = 3  # For Tournament selection
pop = 1000  # Population per generation
gen = 100  # Number of generations
cap = 2500
S1 = 1000
speed = 70
k = round(speed / 60)
price_per_litre = 23  # 1 litre mazot 23 tl
km_per_litre = 5  # 1 litre mazotla 5 km yol kat edilir
tank = 400  # 400 litre mazot deposu
tk = km_per_litre*tank #2000 km

n_list = np.array([rd.sample(X1, len(X1)) for i in range(pop)])
print("population",n_list)

def distance(array):
    path1 = [0]
    total1 = 0  # distance
    total2 = 0  # milk
    X0 = array.copy()
    Xtemp = []
    for f in X0:
         Xtemp.append(int(f))
    # print(".......................................................................................................")
    #print(Xtemp)
    total_distance = 0
    for z in range(S1):
        last = path1[-1]
        next_loc = Xtemp[0]
        Xtemp.remove(next_loc)
        total1 += dist(last, next_loc)
        total2 += milk(next_loc - 1)
        if total2 <= cap and total1 <= tk:
            path1.append(next_loc)
            # print("Minimum Route:", path1, "Total Distance:", total1, "current distance:",dist(last, next_loc), "current milk:", milk(next_loc - 1), ",", "Total Milk:", total2)
        else:
            Xtemp.insert(0, next_loc)
            total2 -= milk(next_loc - 1)
            total1 -= dist(last, next_loc)
            path1.append(0)
            tdk = dist(last, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk  # distance between last station to center
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            # print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            # print("Average turnaround time:", round(((k * total1) / 60)), "hr.")
            # print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            # print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            # print("...................")
        if len(Xtemp) == 0:
            path1.append(0)
            tdk = dist(next_loc, 0)
            total1 += tdk
            for s in range(S1):
                if (total1 > tk):
                    total1 -= dist(path1[len(path1) - 2], 0)
                    total1 -= dist(path1[len(path1) - 3], path1[len(path1) - 2])
                    tdk = dist(path1[len(path1) - 3], 0)
                    total1 += tdk
                    total2 -= milk(path1[len(path1) - 2] - 1)
                    Xtemp.insert(0, path1[len(path1) - 2])
                    path1.remove(path1[len(path1) - 2])
            # print("Minimum Route:", path1, "Total Distance:", total1, "km.", "Total Milk:", total2)
            # print("Average turnaround time:", round(((k * total1) / 60)), "hr.")
            # print("Total amount of fuel consumed:", round(total1 / km_per_litre), "litre")
            # print("Total cost:", round(total1 / km_per_litre) * price_per_litre, "tl")
            total_distance += total1
            total1 = 0
            total2 = 0
            path1 = [0]
            #print("...................")
            if len(Xtemp) == 0:
                 break
    return total_distance

All_in_Generation_X_1 = np.empty((0, len(X1) + 1))  # keep every generations mutant children
All_in_Generation_X_2 = np.empty((0, len(X1) + 1))
Save_Best_in_Generation_X = np.empty((0, len(X1) + 1))  # keep every solution to pick the best one

start_time = time.time()  # start the time
Generation = 1
for i in range(gen):
    print("--> GENERATION: #", Generation)
    Family = 1
    for j in range(int(pop/2)):
        print("--> FAMILY: #", Family)
        #TOURNAMENT SELECTION->decide the 2 parent, the routes with the lowest distance become the parent
        n_list = n_list.astype(int)
        indices = np.random.choice(len(n_list), size=3,replace=False) #choose 3 random solution index,replace=False->produce different indices
        Warrior_1_index = indices[0]
        Warrior_2_index = indices[1]
        Warrior_3_index = indices[2]

        Warrior_1 = n_list[Warrior_1_index, :] #return the solution of the specified index
        Prize_Warrior_1 = distance(Warrior_1) #return distance of the route solution
        Warrior_2 = n_list[Warrior_2_index, :]
        Prize_Warrior_2 = distance(Warrior_2)
        Warrior_3 = n_list[Warrior_3_index, :]
        Prize_Warrior_3 = distance(Warrior_3)

        distances = [Prize_Warrior_1, Prize_Warrior_2, Prize_Warrior_3]
        # Find the two minimum distances
        min_distance_1 = min(distances)
        distances.remove(min_distance_1)
        min_distance_2 = min(distances)

        # Assign the corresponding winners
        if Prize_Warrior_1 == min_distance_1:
            Winner1 = Warrior_1
        elif Prize_Warrior_2 == min_distance_1:
            Winner1 = Warrior_2
        else:
            Winner1 = Warrior_3

        # Assign the corresponding winners
        if Prize_Warrior_1 == min_distance_2:
            Winner2 = Warrior_1
        elif Prize_Warrior_2 == min_distance_2:
            Winner2 = Warrior_2
        else:
            Winner2 = Warrior_3
        print("...............")
        Parent_1 = Winner1
        print("Parent 1",Parent_1)
        Parent_2 = Winner2
        print("Parent 2",Parent_2)

        #CROSSOVER
        Child_1 = np.zeros(len(X1), dtype=int)
        Child_2 = np.zeros(len(X1), dtype=int)

        cros_rnd = rd.random() #choose random number between 0-1 to decide will the crossover happen or not
        if cros_rnd < p_c: #if crossover prob. is bigger crossover happens
            Cr_1 = np.random.randint(0, len(X1)) #to select points to crossover
            Cr_2 = np.random.randint(0, len(X1)) #to select points to crossover
            while Cr_1 == Cr_2:
                Cr_2 = np.random.randint(0, len(X1))

            if Cr_1 < Cr_2: # if second point is bigger
                Cr_2 = Cr_2 + 1 #[1:3] to include 3
                New_Dep_1 = Parent_1[Cr_1:Cr_2]  #gets the location range in parent1
                print("range in parent1",New_Dep_1)
                New_Dep_2 = Parent_2[Cr_1:Cr_2]  #gets the location range in parent2
                print("range in parent2",New_Dep_2)

                #for child 2
                Child_2 = np.concatenate((Child_2[:Cr_1], New_Dep_1, Child_2[Cr_2:]))
                print("Child2",Child_2)

                result = [i for i in Parent_2 if i not in Child_2] #Retrieves missing locations from parent 2 in order
                print("result2",result)

                x = 0
                for i in range(len(Child_2)):
                    if Child_2[i] == 0:
                        Child_2[i] = result[x]
                        x += 1
                print("Child2", Child_2)
                print("...............")

                #for child 1
                Child_1 = np.concatenate((Child_1[:Cr_1], New_Dep_2, Child_1[Cr_2:]))
                print("Child1",Child_1)

                result = [i for i in Parent_1 if i not in Child_1] #Retrieves missing locations from parent 1 in order
                print("result1",result)

                x = 0
                for i in range(len(Child_1)):
                    if Child_1[i] == 0:
                        Child_1[i] = result[x]
                        x += 1
                print("Child1", Child_1)
                print("...............")

            else:  #if first point is bigger
                Cr_1 = Cr_1 + 1 #[1:3] to include 3
                New_Dep_1 = Parent_1[Cr_2:Cr_1]  ##gets the location range in parent1
                print("range in parent1",New_Dep_1)
                New_Dep_2 = Parent_2[Cr_2:Cr_1]  ##gets the location range in parent2
                print("range in parent2",New_Dep_2)

                #for child 2
                Child_2 = np.concatenate((Child_2[:Cr_2], New_Dep_1, Child_2[Cr_1:]))
                print("Child2",Child_2)

                result = [i for i in Parent_2 if i not in Child_2]
                print("result2",result)

                x = 0
                for i in range(len(Child_2)):
                    if Child_2[i] == 0:
                        Child_2[i] = result[x]
                        x += 1
                print("Child2", Child_2)
                print("...............")

                #for child 1
                Child_1 = np.concatenate((Child_1[:Cr_2], New_Dep_2, Child_1[Cr_1:]))
                print("Child1",Child_1)

                result = [i for i in Parent_1 if i not in Child_1]
                print("result1",result)

                x = 0
                for i in range(len(Child_1)):
                    if Child_1[i] == 0:
                        Child_1[i] = result[x]
                        x += 1
                print("Child1", Child_1)
                print("...............")

        else:  #no crossover happen
            Child_1 = Parent_1 #child 1 will be same as parent 1
            Child_2 = Parent_2 #child 2 will be same as parent 2

        #MUTATION
        #For Child1
        Ran_Mut_1 = np.random.rand() #choose random number between 0-1 to decide will the mutation happen or not
        if Ran_Mut_1 < p_m: #if mutation prob. is bigger, mutation happens
            Ran_Mut_2 = np.random.randint(0, len(X1))  # to select points to mutate
            Ran_Mut_3 = np.random.randint(0, len(X1))  # to select points to crossover
            A1 = Ran_Mut_2
            A2 = Ran_Mut_3
            while A1 == A2:
                A2 = np.random.randint(0, len(X1))

            if A1 < A2: #if second index is bigger than first index
                A2 = A2 + 1 #[1:3] to include 3
                Rev_1 = Child_1[:]
                Rev_2 = list(reversed(Child_1[A1:A2])) #reversed the range
                t = 0
                for i in range(A1, A2):
                    Rev_1[i] = Rev_2[t]  # The reversed will take place
                    t += 1
                Mutated_Child_1 = Rev_1

            else: #if first index is bigger than second index
                A1 = A1 + 1 #[1:3] to include 3
                Rev_1 = Child_1[:]
                Rev_2 = list(reversed(Child_1[A2:A1])) #reversed the range
                t = 0
                for i in range(A2, A1):
                    Rev_1[i] = Rev_2[t] # The reversed will take place
                    t = t + 1
                Mutated_Child_1 = Rev_1

        else:
            Mutated_Child_1 = Child_1
        Total_Dist_Mut_1 = distance(Mutated_Child_1)

        #For Child2
        Ran_Mut_1 = rd.random()  # Probablity to Mutate
        if Ran_Mut_1 < p_m:  #if mutation prob. is bigger mutation happens
            Ran_Mut_2 = np.random.randint(0, len(X1))  # to select points to mutate
            Ran_Mut_3 = np.random.randint(0, len(X1))  # to select points to crossover
            A1 = Ran_Mut_2
            A2 = Ran_Mut_3
            while A1 == A2:
                A2 = np.random.randint(0, len(X1))

            if A1 < A2: #if second index is bigger than first index
                A2 = A2 + 1 #[1:3] to include 3
                Rev_1 = Child_2[:]
                Rev_2 = list(reversed(Child_2[A1:A2])) #reversed the range
                t = 0
                for i in range(A1, A2):
                    Rev_1[i] = Rev_2[t] # The reversed will take place
                    t = t + 1
                Mutated_Child_2 = Rev_1

            else: #if first index is bigger than second index
                A1 = A1 + 1 #[1:3] to include 3
                Rev_1 = Child_2[:]
                Rev_2 = list(reversed(Child_2[A2:A1])) #reversed the range
                t = 0
                for i in range(A2, A1):
                    Rev_1[i] = Rev_2[t] # The reversed will take place
                    t = t + 1
                Mutated_Child_2 = Rev_1

        else:
            Mutated_Child_2 = Child_2
        Total_Dist_Mut_2 = distance(Mutated_Child_2)

        print("Mutated child1",Mutated_Child_1)
        new_arr = np.insert(Mutated_Child_1, 0, Total_Dist_Mut_1)
        print("new_arr",new_arr)
        All_in_Generation_X_1 = np.vstack((All_in_Generation_X_1, new_arr.astype(str)))
        print("All_in_Generation_X_1",All_in_Generation_X_1)

        print("Mutated_Child_2",Mutated_Child_2)
        new_arr2 = np.insert(Mutated_Child_2, 0, Total_Dist_Mut_2)
        print("new_arr2",new_arr2)
        All_in_Generation_X_2 = np.vstack((All_in_Generation_X_2, new_arr2.astype(str)))
        print("All_in_Generation_X_2",All_in_Generation_X_2)

        Save_Best_in_Generation_X = np.vstack((All_in_Generation_X_1, All_in_Generation_X_2))
        print("Save_Best_in_Generation_X",Save_Best_in_Generation_X) #saves mutated children1 and mutated children2[['3346' '2' '7' '9' '4' '5' '8' '3' '6' '1' '10']['4053' '3' '8' '6' '4' '9' '7' '1' '5' '10' '2']]

        Family = Family + 1

    n_list = Save_Best_in_Generation_X[:, 1:]
    print("n-list",n_list)

    t = 0
    R_Final = []
    #best
    for i in Save_Best_in_Generation_X:
        if (Save_Best_in_Generation_X[t, :1]) <= min(Save_Best_in_Generation_X[:, :1]):
            R_Final = Save_Best_in_Generation_X[t, :]
        t = t + 1
    print("Best",R_Final)
    #worst
    t = 0
    R_22_Final = []
    for i in Save_Best_in_Generation_X:
        if (Save_Best_in_Generation_X[t, :1]) >= max(Save_Best_in_Generation_X[:, :1]):
            R_22_Final = Save_Best_in_Generation_X[t, :]
        t = t + 1
    print("Final_Worst",R_22_Final)
    Worst_1 = np.where((Save_Best_in_Generation_X == R_22_Final).all(axis=1))
    print("worst route index",Worst_1)

    ch = R_Final[1:]
    print("Final_Best",ch)
    #change the worst one with the best one
    n_list[Worst_1] =ch
    print("New Population",n_list)

    Generation = Generation + 1
print("Solution",ch)
comp_time = time.time() - start_time  # Keeps the difference between the end time and the start time
print(f"-> Computational Time: {comp_time} seconds")