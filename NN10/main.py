import numpy as np
from openpyxl import Workbook, load_workbook
import time

wb = load_workbook("dist1.xlsx")
ws = wb.active
wb1 = load_workbook("milk1.xlsx")
ws1 = wb1.active
A = []
B = []
C = []
for x in range(1, 12):
    for y in range(1, 12):
        A.append(round(ws.cell(x, y).value))
x = np.array(A)
shape = (11, 11)
A = x.reshape(shape)
#print(A)


for x in range(1, 11):
    for y in range(1, 2):
        B.append(ws1.cell(x, y).value)

def dist(fr, to):
    return A[fr][to]

def milk(fr):
    return B[fr]



cap = 2500
speed = 70
k = round(speed / 60) #1
price_per_litre = 23  # 1 litre mazot 23 tl
km_per_litre = 5  # 1 litre mazot 5 km yol kat edilir
tank = 400  # 400 litre mazot deposu
S = 100
tk = km_per_litre*tank #2000 km
start_time = time.time()  # Keeps the start time
total_Distance = 0
totalMilk = 0
mask = [True] * len(A)
mask[0] = False
path = [0]
for i in range(S):
        last = path[-1]
        min_val = float('inf')
        for i in range(len(A)):
            if mask[i] and A[last][i] < min_val:
                min_val = A[last][i]
                next_loc = i
        mask[next_loc] = False

        total_Distance += dist(last, next_loc)
        totalMilk += milk(next_loc - 1)

        if (tk - total_Distance >= 0 and cap-totalMilk >= 0):
            path += [next_loc]
            print("Minimum Route:", path, "Total Distance:", total_Distance, "current distance:", dist(last, next_loc), "km.", "Total Milk:", totalMilk)
        else:
            mask[next_loc] = True
            totalMilk -= milk(next_loc-1)
            total_Distance -= dist(last, next_loc)
            path.append(0)
            tdk = dist(last, 0)
            total_Distance += tdk
            for i in range(S):
                if (tk - total_Distance < 0):
                    total_Distance -= dist(path[len(path) - 2], 0)
                    total_Distance -= dist(path[len(path) - 3], path[len(path) - 2])
                    tdk = dist(path[len(path) - 3], 0)
                    total_Distance += tdk #distance between last station to center
                    totalMilk -= milk(path[len(path) - 2]-1)
                    mask[path[len(path) - 2]] = True
                    path.remove(path[len(path) - 2])
            print("Minimum Route:", path,"Total Distance:", total_Distance,"km.","Total Milk:",totalMilk)
            print("Average turnaround time:", round(((k * total_Distance) /60)), "hr.")
            print("Total amount of fuel consumed:", round(total_Distance / km_per_litre), "litre")
            print("Total cost:",round(total_Distance / km_per_litre) * price_per_litre, "tl")
            total_Distance = 0
            totalMilk = 0
            path = [0]
            print("...................")

        all_false = True
        for i in mask:
            if i:
                all_false = False
                break

        if all_false:
            path.append(0)
            tdk = dist(next_loc, 0)
            total_Distance += tdk
            for i in range(S):
                if (tk - total_Distance < 0):
                    total_Distance -= dist(path[len(path) - 2], 0)
                    total_Distance -= dist(path[len(path) - 3], path[len(path) - 2])
                    tdk = dist(path[len(path) - 3], 0)
                    total_Distance += tdk
                    totalMilk -= milk(path[len(path) - 2] - 1)
                    mask[path[len(path) - 2]] = True
                    path.remove(path[len(path) - 2])
            print("Minimum Route:", path, "Total Distance:", total_Distance, "km.", "Total Milk:", totalMilk)
            print("Average turnaround time:", round(((k * total_Distance) / 60)), "hr.")
            print("Total amount of fuel consumed:", round(total_Distance / km_per_litre), "litre")
            print("Total cost:", round(total_Distance / km_per_litre) * price_per_litre, "tl")
            total_Distance = 0
            totalMilk = 0
            path = [0]
            print("...................")
            for i in mask:
                if i:
                    all_false = False
                    break
            if all_false:
                break

comp_time = time.time() - start_time  # Keeps the difference between the end time and the start time
print(f"-> Computational Time: {comp_time} seconds")  # Prints Computational Time
